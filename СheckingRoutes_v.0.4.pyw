import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter.messagebox import showinfo
import openpyxl
import copy
import threading


class ProgressWindow:
    def __init__(self, parent, title="Обработка файла"):
        self.progress_window = tk.Toplevel(parent)
        self.progress_window.title(title)
        self.progress_window.geometry('300x150')
        self.progress_window.transient(parent)
        self.progress_window.grab_set()
        
        self.label = tk.Label(self.progress_window, text="Обработка данных...")
        self.label.pack(pady=10)
        
        self.progress = ttk.Progressbar(
            self.progress_window, 
            orient="horizontal", 
            length=200, 
            mode="determinate"
        )
        self.progress.pack(pady=10)
        
        self.percentage_label = tk.Label(self.progress_window, text="0%")
        self.percentage_label.pack(pady=5)

    def update_progress(self, value, text=None):
        self.progress['value'] = value
        self.percentage_label['text'] = f"{int(value)}%"
        if text:
            self.label['text'] = text

    def close(self):
        self.progress_window.destroy()


# Создаем главное окно
results_window = tk.Tk()
results_window.title("Проверка выгрузки маршрутов")
results_window.geometry('600x500')
results_window.configure(bg='#003366')

# Создаем главный контейнер с отступами
main_frame = ttk.Frame(results_window, padding="20")
main_frame.pack(fill=tk.BOTH, expand=True)

# Создаем стиль для виджетов
style = ttk.Style()
style.configure('Header.TLabel', font=('Arial', 12, 'bold'))
style.configure('Info.TLabel', font=('Arial', 10))
style.configure('Action.TButton', font=('Arial', 10))

# Заголовок
header = ttk.Label(
    main_frame,
    text='Проверка выгрузки маршрутов',
    style='Header.TLabel'
)
header.pack(pady=(0, 20))

# Контейнер для инструкций
instructions_frame = ttk.Frame(main_frame)
instructions_frame.pack(fill=tk.X, pady=(0, 20))

instructions = [
    "1. Для проверки файла нажмите кнопку 'Выбрать файл'",
    "2. После выбора файла начнется проверка наличия всех маршрутов",
    "3. Результат проверки отобразится под кнопкой"
]

for instruction in instructions:
    ttk.Label(
        instructions_frame,
        text=instruction,
        style='Info.TLabel',
        wraplength=550
    ).pack(anchor='w', pady=2)

# Создаем контейнер для кнопок
button_frame = ttk.Frame(main_frame)
button_frame.pack(pady=20)

# Создаем контейнер для результатов
results_frame = ttk.Frame(main_frame)
results_frame.pack(fill=tk.BOTH, expand=True, pady=20)

def add_result_label(text):
    label = ttk.Label(
        results_frame,
        text=text,
        style='Info.TLabel',
        wraplength=550
    )
    label.pack(anchor='w', pady=2)


# Функции обработки файлов
def file_selection():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    if file_path:
        progress = ProgressWindow(results_window, "Проверка файла")
        
        def process_file():
            try:
                progress.update_progress(0, "Загрузка файла...")
                global wb
                wb = openpyxl.load_workbook(file_path)
                ws = wb['Лист1']

                progress.update_progress(33, "Подготовка списка маршрутов...")
                # Встроенный список маршрутов
                check_list = [
                    "10", "101", "101Э", "104/148", "109", "109Б", "110", "112",
                    "120", "121", "122", "125", "126", "134А", "135", "139",
                    "14", "154А", "166", "167", "170", "173", "175", "180А/75А",
                    "182", "194", "199", "1КР", "202", "207", "208", "21",
                    "211", "211", "211Э", "213", "215", "216", "216А", "219",
                    "223", "227", "227", "230", "234", "235", "237", "240",
                    "248", "25", "250", "251", "258", "259", "263", "264",
                    "267", "275", "275", "279", "283", "285", "293", "294",
                    "2КР", "302", "303", "304", "305", "306", "307/309", "310",
                    "311/312", "315.", "318", "319", "320", "321", "322", "38",
                    "399", "3КР", "408", "410", "420", "420А", "420Б", "433",
                    "435", "436", "45", "483", "483А", "494", "497", "552",
                    "404", "567", "60", "69", "75", "84", "78", "79", "84",
                    "85", "99", "462", "531", "557"
                ]

                progress.update_progress(66, "Проверка маршрутов...")
                column_values = [str(cell.value) if cell.value is not None else "" for cell in ws['B']]
                missing_data = [item for item in check_list if item not in column_values]

                progress.update_progress(100, "Завершение проверки...")
                progress.close()

                if missing_data:
                    add_result_label("Не выгружены:")
                    for item in missing_data:
                        add_result_label(item)
                else:
                    showinfo(title="info", message='Все маршруты успешно выгружены.')
                    add_result_label("Все маршруты успешно выгружены.")
                    
            except Exception as e:
                progress.close()
                showinfo("Ошибка", f"Произошла ошибка при проверке файла:\n{str(e)}")

        thread = threading.Thread(target=process_file)
        thread.start()

def copy_cell_format(source_cell, target_cell):
    """Копирование форматирования ячейки с созданием новых объектов стилей"""
    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)


def generate_report_fixed():
    if 'wb' not in globals():
        showinfo(title="Ошибка", message="Сначала выберите файл!")
        return

    ws = wb['Лист1']
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    
    if not output_file:
        return

    progress = ProgressWindow(results_window)
    
    def process_file():
        try:
            new_wb = openpyxl.Workbook()
            # Переименовываем первый лист
            new_ws = new_wb.active
            new_ws.title = "Выбранные маршруты"
            
            # Создаем лист для итоговых показателей
            summary_ws = new_wb.create_sheet("Итоговые показатели")
            
            total_rows = sum(1 for row in ws.iter_rows())
            processed_rows = 0

            # Копируем шапку для первого листа
            header_row = next(ws.iter_rows(max_row=1))
            for col_idx, cell in enumerate(header_row, 1):
                new_cell = new_ws.cell(row=1, column=col_idx, value=cell.value)
                copy_cell_format(cell, new_cell)
            
            processed_rows += 1
            progress.update_progress((processed_rows / total_rows) * 100, "Копирование данных...")

            # Словарь для хранения данных по каждому маршруту
            route_data = {
                '462': {},
                '531': {},
                '557': {}
            }
            
            # Копируем строки с нужными маршрутами и собираем данные для сводной таблицы
            target_routes = ["557", "531", "462"]
            new_row_idx = 2
            
            for row in ws.iter_rows(min_row=2):
                route = str(row[1].value)
                if route in target_routes:
                    # Копируем строку в первый лист
                    for col_idx, cell in enumerate(row, 1):
                        new_cell = new_ws.cell(row=new_row_idx, column=col_idx, value=cell.value)
                        copy_cell_format(cell, new_cell)
                    new_row_idx += 1
                    
                    # Собираем данные для сводной таблицы
                    vehicle_number = str(row[3].value)  # Столбец D - гос. номер ТС
                    passengers = row[9].value or 0  # Столбец J - количество пассажиров
                    
                    if vehicle_number not in route_data[route]:
                        route_data[route][vehicle_number] = 0
                    route_data[route][vehicle_number] += passengers
                
                processed_rows += 1
                if processed_rows % 10 == 0:
                    progress.update_progress((processed_rows / total_rows) * 100)

            # Создаем сводную таблицу на втором листе
            progress.update_progress(90, "Формирование сводной таблицы...")
            
            # Заголовки для сводной таблицы
            summary_ws['A1'] = 'Маршрут'
            summary_ws['B1'] = 'Гос. номер ТС'
            summary_ws['C1'] = 'Количество перевезенных пассажиров'
            
            # Применяем стили к заголовкам
            for cell in summary_ws['1:1']:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Заполняем данные
            current_row = 2
            for route in target_routes:
                first_row = current_row
                for vehicle, passengers in route_data[route].items():
                    summary_ws[f'A{current_row}'] = route
                    summary_ws[f'B{current_row}'] = vehicle
                    summary_ws[f'C{current_row}'] = passengers
                    current_row += 1
                
                # Добавляем итоговую строку для маршрута
                summary_ws[f'A{current_row}'] = f'Итого по маршруту {route}'
                summary_ws[f'C{current_row}'] = f'=SUM(C{first_row}:C{current_row-1})'
                summary_ws[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
                summary_ws[f'C{current_row}'].font = openpyxl.styles.Font(bold=True)
                current_row += 1
                
                # Добавляем пустую строку между маршрутами
                current_row += 1

            # Устанавливаем ширину столбцов
            summary_ws.column_dimensions['A'].width = 15
            summary_ws.column_dimensions['B'].width = 20
            summary_ws.column_dimensions['C'].width = 40

            # Копируем форматирование для первого листа
            progress.update_progress(95, "Копирование форматирования...")
            for column in ws.columns:
                col_letter = openpyxl.utils.get_column_letter(column[0].column)
                if col_letter in ws.column_dimensions:
                    source_column_width = ws.column_dimensions[col_letter].width
                    if not source_column_width:
                        source_column_width = 8.43
                    new_ws.column_dimensions[col_letter].width = source_column_width

            progress.update_progress(98, "Сохранение файла...")
            new_wb.save(output_file)
            
            progress.close()
            showinfo("Успех", "Файл успешно обработан и сохранен!")
            
        except Exception as e:
            progress.close()
            showinfo("Ошибка", f"Произошла ошибка при обработке файла:\n{str(e)}")

    thread = threading.Thread(target=process_file)
    thread.start()

def remove_routes():
    if 'wb' not in globals():
        showinfo(title="Ошибка", message="Сначала выберите файл!")
        return
    
    # Сразу запрашиваем путь для нового файла
    output_file = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Сохранить результат как"
    )
    
    if not output_file:  # Если пользователь отменил выбор файла
        return
    
    progress = ProgressWindow(results_window, "Удаление маршрутов")
    
    def process_deletion():
        try:
            ws = wb['Лист1']
            progress.update_progress(0, "Обработка данных...")
            
            # Получаем все данные из листа
            data = list(ws.values)
            if not data:
                progress.close()
                showinfo("Ошибка", "Файл пуст!")
                return
                
            # Создаем новый Excel-файл
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            
            # Копируем шапку с форматированием
            header_row = next(ws.iter_rows(max_row=1))
            for col_idx, cell in enumerate(header_row, 1):
                new_cell = new_ws.cell(row=1, column=col_idx, value=cell.value)
                copy_cell_format(cell, new_cell)

            for column in ws.columns:
                col_letter = openpyxl.utils.get_column_letter(column[0].column)
                if col_letter in ws.column_dimensions:
                    source_width = ws.column_dimensions[col_letter].width
                    new_ws.column_dimensions[col_letter].width = source_width if source_width else 8.43

            # Получаем заголовки
            headers = data[0]
            
            # Фильтруем данные, исключая указанные маршруты
            routes_to_delete = {"557", "531", "462"}
            filtered_data = [headers]  # Сохраняем заголовки
            rows_filtered = 0
            total_rows = len(data) - 1
            
            progress.update_progress(20, "Фильтрация данных...")
            
            for row in data[1:]:
                if str(row[1]) not in routes_to_delete:
                    filtered_data.append(row)
                rows_filtered += 1
                if rows_filtered % 1000 == 0:
                    progress.update_progress(20 + (rows_filtered / total_rows * 60))
            
            deleted_count = total_rows - (len(filtered_data) - 1)
            
            progress.update_progress(90, "Запись данных в новый файл...")
            for row_idx, row_data in enumerate(filtered_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    new_ws.cell(row=row_idx, column=col_idx, value=value)
            
            progress.update_progress(95, "Сохранение файла...")
            new_wb.save(output_file)
            
            progress.close()
            showinfo("Успех", 
                    f"Маршруты 557, 531 и 462 успешно удалены!\n"
                    f"Удалено строк: {deleted_count}\n"
                    f"Результат сохранен в новый файл")
            
        except Exception as e:
            progress.close()
            showinfo("Ошибка", f"Произошла ошибка при удалении маршрутов:\n{str(e)}")

    thread = threading.Thread(target=process_deletion)
    thread.start()


# Кнопки
select_button = ttk.Button(
    button_frame,
    text="Выбрать файл",
    style='Action.TButton',
    command=file_selection,
    width=40
)
select_button.pack(pady=(0, 10))

report_button = ttk.Button(
    button_frame,
    text="Сформировать отчёт по маршрутам № 557, 531, 462",
    style='Action.TButton',
    command=generate_report_fixed,
    width=50
)
report_button.pack(pady=(0, 10))

remove_button = ttk.Button(
    button_frame,
    text="Удалить маршруты 557, 531 и 462 из файла",
    style='Action.TButton',
    command=remove_routes,
    width=40
)
remove_button.pack(pady=(0, 10))

results_window.mainloop()